# core/export.py
"""Export land records to Excel and PDF."""
import os
import json
import datetime
from typing import List


def export_to_excel(records: List[dict], output_path: str) -> str:
    """Export list of land records to Excel file. Returns output_path."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # -----------------------------------------------------------------------
    # Sheet 1: GCN Records
    # -----------------------------------------------------------------------
    ws_gcn = wb.create_sheet("GCN - Sổ đỏ")
    gcn_headers = [
        "STT", "Số GCN", "Ngày cấp", "Cơ quan cấp",
        "Chủ sở hữu", "CMND/CCCD", "Ngày sinh", "Địa chỉ chủ",
        "Số tờ BĐ", "Số thửa", "Diện tích (m²)", "Mục đích SD",
        "Thời hạn SD", "Địa chỉ thửa đất",
        "Phường/Xã", "Quận/Huyện", "Tỉnh/Thành",
        "Trạng thái", "Ngày nhập"
    ]
    gcn_records = [r for r in records if r.get("doc_type") == "gcn"]
    hop_dong_records = [r for r in records if r.get("doc_type") == "hop_dong"]

    header_fill = PatternFill("solid", fgColor="1a3a6b")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def _style_header(ws, headers):
        ws.append(headers)
        for col, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[1].height = 30

    _style_header(ws_gcn, gcn_headers)

    for stt, rec in enumerate(gcn_records, 1):
        raw = json.loads(rec.get("raw_ocr", "{}")) if isinstance(rec.get("raw_ocr"), str) else {}
        owners = raw.get("chu_so_huu", [])
        owner = owners[0] if owners else {}
        detail = rec.get("detail", raw)

        row = [
            stt,
            detail.get("so_gcn", raw.get("so_gcn", "")),
            detail.get("ngay_cap", raw.get("ngay_cap", "")),
            detail.get("co_quan_cap", raw.get("co_quan_cap", "")),
            owner.get("ho_ten", "") if isinstance(owner, dict) else str(owner),
            owner.get("cmnd_cccd", "") if isinstance(owner, dict) else "",
            owner.get("ngay_sinh", "") if isinstance(owner, dict) else "",
            owner.get("dia_chi", "") if isinstance(owner, dict) else "",
            detail.get("so_to_ban_do", raw.get("so_to_ban_do", "")),
            detail.get("so_thua_dat", raw.get("so_thua_dat", "")),
            detail.get("dien_tich", raw.get("dien_tich", "")),
            detail.get("muc_dich_su_dung", raw.get("muc_dich_su_dung", "")),
            detail.get("thoi_han_su_dung", raw.get("thoi_han_su_dung", "")),
            detail.get("dia_chi_thua_dat", raw.get("dia_chi_thua_dat", "")),
            detail.get("phuong_xa", raw.get("phuong_xa", "")),
            detail.get("quan_huyen", raw.get("quan_huyen", "")),
            detail.get("tinh_thanh", raw.get("tinh_thanh", "")),
            rec.get("status", "draft"),
            rec.get("created_at", "")
        ]
        ws_gcn.append(row)
        row_idx = stt + 1
        for col in range(1, len(gcn_headers) + 1):
            cell = ws_gcn.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if stt % 2 == 0:
                cell.fill = alt_fill

    # Auto-fit column widths
    for col in ws_gcn.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws_gcn.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)
    ws_gcn.freeze_panes = "A2"

    # -----------------------------------------------------------------------
    # Sheet 2: Hop Dong Records
    # -----------------------------------------------------------------------
    ws_hd = wb.create_sheet("Hợp đồng chuyển nhượng")
    hd_headers = [
        "STT", "Số HĐ", "Ngày ký", "Nơi ký",
        "Bên chuyển nhượng", "CMND/CCCD",
        "Bên nhận chuyển nhượng", "CMND/CCCD",
        "Số tờ BĐ", "Số thửa", "Diện tích (m²)",
        "Địa chỉ thửa đất", "Số GCN liên quan",
        "Giá chuyển nhượng (VNĐ)", "Phương thức TT",
        "Công chứng viên", "VP công chứng", "Số CC",
        "Ngày CC", "Trạng thái", "Ngày nhập"
    ]
    _style_header(ws_hd, hd_headers)

    for stt, rec in enumerate(hop_dong_records, 1):
        raw = json.loads(rec.get("raw_ocr", "{}")) if isinstance(rec.get("raw_ocr"), str) else {}
        ben_chuyen = (raw.get("ben_chuyen_nhuong") or [{}])[0] if isinstance(raw.get("ben_chuyen_nhuong"), list) else {}
        ben_nhan = (raw.get("ben_nhan_chuyen_nhuong") or [{}])[0] if isinstance(raw.get("ben_nhan_chuyen_nhuong"), list) else {}
        thua = raw.get("thong_tin_thua_dat") or {}
        detail = rec.get("detail", raw)

        row = [
            stt,
            detail.get("so_hop_dong", raw.get("so_hop_dong", "")),
            detail.get("ngay_ky", raw.get("ngay_ky", "")),
            detail.get("noi_ky", raw.get("noi_ky", "")),
            ben_chuyen.get("ho_ten", "") if isinstance(ben_chuyen, dict) else "",
            ben_chuyen.get("cmnd_cccd", "") if isinstance(ben_chuyen, dict) else "",
            ben_nhan.get("ho_ten", "") if isinstance(ben_nhan, dict) else "",
            ben_nhan.get("cmnd_cccd", "") if isinstance(ben_nhan, dict) else "",
            thua.get("so_to_ban_do", detail.get("so_to_ban_do", "")),
            thua.get("so_thua_dat", detail.get("so_thua_dat", "")),
            thua.get("dien_tich", detail.get("dien_tich", "")),
            thua.get("dia_chi", detail.get("dia_chi_thua_dat", "")),
            thua.get("so_gcn", detail.get("so_gcn_lien_quan", "")),
            detail.get("gia_chuyen_nhuong", raw.get("gia_chuyen_nhuong", "")),
            detail.get("phuong_thuc_thanh_toan", raw.get("phuong_thuc_thanh_toan", "")),
            detail.get("cong_chung_vien", raw.get("cong_chung_vien", "")),
            detail.get("van_phong_cong_chung", raw.get("van_phong_cong_chung", "")),
            detail.get("so_cong_chung", raw.get("so_cong_chung", "")),
            detail.get("ngay_cong_chung", raw.get("ngay_cong_chung", "")),
            rec.get("status", "draft"),
            rec.get("created_at", "")
        ]
        ws_hd.append(row)
        row_idx = stt + 1
        for col in range(1, len(hd_headers) + 1):
            cell = ws_hd.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if stt % 2 == 0:
                cell.fill = alt_fill

    for col in ws_hd.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws_hd.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)
    ws_hd.freeze_panes = "A2"

    wb.save(output_path)
    return output_path


def export_record_to_pdf(record: dict, output_path: str) -> str:
    """Export a single land record to a formatted PDF."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable)
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import json

    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            topMargin=2*cm, bottomMargin=2*cm,
                            leftMargin=2.5*cm, rightMargin=2*cm)

    styles = getSampleStyleSheet()
    NAVY = colors.HexColor("#1a3a6b")
    LIGHT_BLUE = colors.HexColor("#EBF3FB")
    ACCENT = colors.HexColor("#2563eb")

    title_style = ParagraphStyle("Title", parent=styles["Title"],
                                  fontSize=16, textColor=NAVY, spaceAfter=6)
    heading_style = ParagraphStyle("Heading", parent=styles["Heading2"],
                                    fontSize=11, textColor=ACCENT, spaceBefore=12, spaceAfter=4)
    body_style = ParagraphStyle("Body", parent=styles["Normal"],
                                 fontSize=9.5, leading=14)
    label_style = ParagraphStyle("Label", parent=styles["Normal"],
                                  fontSize=9, textColor=colors.grey)

    raw = json.loads(record.get("raw_ocr", "{}")) if isinstance(record.get("raw_ocr"), str) else {}
    detail = record.get("detail", raw)
    doc_type = record.get("doc_type", "generic")
    owners = record.get("owners", [])

    story = []
    now_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")

    # Header
    title = "GIẤY CHỨNG NHẬN QSDĐ" if doc_type == "gcn" else "HỢP ĐỒNG CHUYỂN NHƯỢNG"
    story.append(Paragraph(f"PHIẾU SỐ HÓA: {title}", title_style))
    story.append(Paragraph(f"Ngày in: {now_str} | Trạng thái: {record.get('status','draft').upper()}", label_style))
    story.append(HRFlowable(width="100%", thickness=2, color=NAVY, spaceAfter=12))

    def _row(label, val):
        return [Paragraph(f"<b>{label}</b>", body_style), Paragraph(str(val) if val else "—", body_style)]

    if doc_type == "gcn":
        story.append(Paragraph("I. THÔNG TIN GIẤY CHỨNG NHẬN", heading_style))
        tbl_data = [
            _row("Số GCN:", detail.get("so_gcn", "")),
            _row("Ngày cấp:", detail.get("ngay_cap", "")),
            _row("Cơ quan cấp:", detail.get("co_quan_cap", "")),
        ]
        story.append(Table(tbl_data, colWidths=[5*cm, 12*cm],
                           style=TableStyle([
                               ("BACKGROUND", (0,0), (0,-1), LIGHT_BLUE),
                               ("GRID", (0,0), (-1,-1), 0.5, colors.lightgrey),
                               ("VALIGN", (0,0), (-1,-1), "TOP"),
                               ("TOPPADDING", (0,0), (-1,-1), 4),
                           ])))

        story.append(Paragraph("II. THÔNG TIN CHỦ SỞ HỮU", heading_style))
        for owner in owners:
            tbl_data = [
                _row("Họ và tên:", owner.get("ho_ten", "")),
                _row("Ngày sinh:", owner.get("ngay_sinh", "")),
                _row("CMND/CCCD:", owner.get("cmnd_cccd", "")),
                _row("Địa chỉ:", owner.get("dia_chi", "")),
            ]
            story.append(Table(tbl_data, colWidths=[5*cm, 12*cm],
                               style=TableStyle([
                                   ("BACKGROUND", (0,0), (0,-1), LIGHT_BLUE),
                                   ("GRID", (0,0), (-1,-1), 0.5, colors.lightgrey),
                                   ("VALIGN", (0,0), (-1,-1), "TOP"),
                                   ("TOPPADDING", (0,0), (-1,-1), 4),
                               ])))

        story.append(Paragraph("III. THÔNG TIN THỬA ĐẤT", heading_style))
        tbl_data = [
            _row("Số tờ bản đồ:", detail.get("so_to_ban_do", "")),
            _row("Số thửa:", detail.get("so_thua_dat", "")),
            _row("Diện tích:", f"{detail.get('dien_tich', '')} {detail.get('dien_tich_don_vi','m²')}"),
            _row("Mục đích sử dụng:", detail.get("muc_dich_su_dung", "")),
            _row("Thời hạn sử dụng:", detail.get("thoi_han_su_dung", "")),
            _row("Nguồn gốc SD:", detail.get("nguon_goc_su_dung", "")),
            _row("Địa chỉ thửa đất:", detail.get("dia_chi_thua_dat", "")),
            _row("Phường/Xã:", detail.get("phuong_xa", "")),
            _row("Quận/Huyện:", detail.get("quan_huyen", "")),
            _row("Tỉnh/Thành:", detail.get("tinh_thanh", "")),
        ]
        story.append(Table(tbl_data, colWidths=[5*cm, 12*cm],
                           style=TableStyle([
                               ("BACKGROUND", (0,0), (0,-1), LIGHT_BLUE),
                               ("GRID", (0,0), (-1,-1), 0.5, colors.lightgrey),
                               ("VALIGN", (0,0), (-1,-1), "TOP"),
                               ("TOPPADDING", (0,0), (-1,-1), 4),
                           ])))

    elif doc_type == "hop_dong":
        story.append(Paragraph("I. THÔNG TIN HỢP ĐỒNG", heading_style))
        tbl_data = [
            _row("Số hợp đồng:", detail.get("so_hop_dong", "")),
            _row("Ngày ký:", detail.get("ngay_ky", "")),
            _row("Nơi ký:", detail.get("noi_ky", "")),
            _row("Giá chuyển nhượng:", f"{detail.get('gia_chuyen_nhuong','')} {detail.get('don_vi_tien','VNĐ')}"),
            _row("Phương thức TT:", detail.get("phuong_thuc_thanh_toan", "")),
        ]
        story.append(Table(tbl_data, colWidths=[5*cm, 12*cm],
                           style=TableStyle([
                               ("BACKGROUND", (0,0), (0,-1), LIGHT_BLUE),
                               ("GRID", (0,0), (-1,-1), 0.5, colors.lightgrey),
                               ("VALIGN", (0,0), (-1,-1), "TOP"),
                               ("TOPPADDING", (0,0), (-1,-1), 4),
                           ])))

        story.append(Paragraph("II. CÁC BÊN", heading_style))
        for owner in owners:
            role_label = "Bên chuyển nhượng" if owner.get("role") == "ben_chuyen" else "Bên nhận chuyển nhượng"
            tbl_data = [
                _row(f"{role_label} - Họ tên:", owner.get("ho_ten", "")),
                _row("CMND/CCCD:", owner.get("cmnd_cccd", "")),
                _row("Ngày sinh:", owner.get("ngay_sinh", "")),
                _row("Địa chỉ:", owner.get("dia_chi", "")),
            ]
            story.append(Table(tbl_data, colWidths=[5*cm, 12*cm],
                               style=TableStyle([
                                   ("BACKGROUND", (0,0), (0,-1), LIGHT_BLUE),
                                   ("GRID", (0,0), (-1,-1), 0.5, colors.lightgrey),
                                   ("VALIGN", (0,0), (-1,-1), "TOP"),
                                   ("TOPPADDING", (0,0), (-1,-1), 4),
                               ])))

        story.append(Paragraph("III. THÔNG TIN THỬA ĐẤT", heading_style))
        tbl_data = [
            _row("Số tờ BĐ:", detail.get("so_to_ban_do", "")),
            _row("Số thửa:", detail.get("so_thua_dat", "")),
            _row("Diện tích:", f"{detail.get('dien_tich','')} m²"),
            _row("Địa chỉ:", detail.get("dia_chi_thua_dat", "")),
            _row("Số GCN liên quan:", detail.get("so_gcn_lien_quan", "")),
        ]
        story.append(Table(tbl_data, colWidths=[5*cm, 12*cm],
                           style=TableStyle([
                               ("BACKGROUND", (0,0), (0,-1), LIGHT_BLUE),
                               ("GRID", (0,0), (-1,-1), 0.5, colors.lightgrey),
                               ("VALIGN", (0,0), (-1,-1), "TOP"),
                               ("TOPPADDING", (0,0), (-1,-1), 4),
                           ])))

        story.append(Paragraph("IV. CÔNG CHỨNG", heading_style))
        tbl_data = [
            _row("Công chứng viên:", detail.get("cong_chung_vien", "")),
            _row("VP công chứng:", detail.get("van_phong_cong_chung", "")),
            _row("Số công chứng:", detail.get("so_cong_chung", "")),
            _row("Ngày công chứng:", detail.get("ngay_cong_chung", "")),
        ]
        story.append(Table(tbl_data, colWidths=[5*cm, 12*cm],
                           style=TableStyle([
                               ("BACKGROUND", (0,0), (0,-1), LIGHT_BLUE),
                               ("GRID", (0,0), (-1,-1), 0.5, colors.lightgrey),
                               ("VALIGN", (0,0), (-1,-1), "TOP"),
                               ("TOPPADDING", (0,0), (-1,-1), 4),
                           ])))

    # Footer
    story.append(Spacer(1, 1*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.lightgrey))
    story.append(Paragraph(f"File nguồn: {record.get('source_file','N/A')} | Ngày nhập: {record.get('created_at','')}", label_style))

    doc.build(story)
    return output_path
